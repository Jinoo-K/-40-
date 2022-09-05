# 네이버 메일을 발신자로 html 메일 발송
from openpyxl import load_workbook
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart

load_wb = load_workbook(r'14. 구글 및 네이버 이메일 보내기 및 대량 이메일 전송\이메일주소.xlsx', data_only=True)
load_ws = load_wb.active

for i in range (1, load_ws.max_row + 1) : 
    recv_email_value = load_ws.cell(i, 1).value
    print("성공:", recv_email_value)
    try : 
        send_email = "네이버 메일 주소"
        send_pwd = "네이버 메일 비밀번호"

        recv_email = recv_email_value

        smtp_name = "smtp.naver.com"
        smtp_port = 587

        msg = MIMEMultipart()

        msg['Subject'] = "엑셀에서 메일 주소를 읽어 자동으로 보내는 메일입니다.."
        msg['From'] = send_email
        msg['To'] = recv_email

        text = """
        메일 내용입니다.
        """

        msg.attach(MIMEText(text))

        s = smtplib.SMTP(smtp_name, smtp_port)      # 메일 발송
        s.starttls()
        s.login(send_email, send_pwd)
        s.sendmail(send_email, recv_email, msg.as_string())
        s.quit
    except :
        print("에러:",recv_email_value)