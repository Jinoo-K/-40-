import pandas as pd
import requests
from openpyxl import load_workbook
from openpyxl import Workbook
import re

filePath = r"27. 전국의 대학교 위치 시각화하기\고등교육기관 주소록(2022)_학교코드 포함.xlsx"
df_from_excel = pd.read_excel(filePath, engine="openpyxl")
df_from_excel.columns = df_from_excel.loc[4].tolist()
df_from_excel = df_from_excel.drop(index=list(range(0, 5)))

url = 'http://api.vworld.kr/req/address?'
params = 'service=address&request=getcoord&version=2.0&crs=epsg:4326&refine=true&simple=false&format=json&type='
road_type = 'ROAD'
road_type2 = 'PARCEL'
address = '&address='
keys = '&key='
primary_key = ''

def request_geo(road):
    page = requests.get(url+params+road_type+address+road+keys+primary_key)
    json_data = page.json()
    if json_data['response']['status'] == 'OK' :
        x = json_data['response']['result']['point']['x']
        y = json_data['response']['result']['point']['y']
        return x, y
    else : 
        x = 0
        y = 0
        return x, y
    
try :
    wb = load_workbook(r"27. 전국의 대학교 위치 시각화하기\학교주소좌표.xlsx", data_only=True)
    sheet = wb.active
except :
    wb = Workbook()
    sheet = wb.active

university_list = df_from_excel["학교명"].to_list()
address_list = df_from_excel["주소"].to_list()

for num, value in enumerate(address_list) :
    addr = re.sub(r'\([^)]*\)', '', value)
    print(addr)
    x, y = request_geo(addr)
    sheet.append([university_list[num], addr, x, y])
    
wb.save(r"27. 전국의 대학교 위치 시각화하기\학교주소좌표.xlsx")

